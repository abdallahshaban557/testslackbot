import snowflake.connector
import os
#needed to get the environment variables
from dotenv import load_dotenv
from os.path import join, dirname
#excel library
from openpyxl import Workbook, load_workbook

#Environment file
dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

#snowflake
ctx = snowflake.connector.connect(
    user = os.environ.get('snowflake_username'),
    password=os.environ.get('snowflake_password'),
    account='petco.us-east-1'
        )
cs = ctx.cursor()
try:
    cs.execute("SELECT current_version()")
    one_row = cs.fetchone()
    print(one_row[0])
finally:
    cs.close()




def BOPUS_METRICS():
#Query to get snowflake data
        query = ctx.cursor()
        query_file = open('./queries/Fill rate and other BOPUS metrics.sql')               
        content_of_query = query_file.read()
        #query.execute('SELECT SHIPMENT_KEY,ORDER_HEADER_KEY FROM "WHPRD_VW"."DWADMIN"."F_OMS_YFS_SHIPMENT" LIMIT 10')
        query.execute(content_of_query)
        one_row = query.fetchall()
        description = query.description
        number_of_columns = len(description)
        #get length of rows from snowflake array
        query_length = len(one_row)
        #create excel workbook
        wb = Workbook()
        ws = wb.active
        for j in range (0,number_of_columns):
            ws.cell(row=1, column=j+1).value = description[j][0]
        for i in range (0,query_length):
            #get the active worksheet
            for j in range(0, 5):
                #print(one_row[i][j])
                ws.cell(row=i+2, column=j+1).value = one_row[i][j]
            
        #saves query file
        wb.save('query_results\BOPUS_Metrics.xlsx')
        query.close()